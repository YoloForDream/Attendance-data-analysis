#-*- encoding: utf-8 -*- 
import time
import paramiko
import datetime
import timedelta
import openpyxl
import csv
import os
import pandas as pd
from pandas.tseries.offsets import Day
from pandas import read_csv
import os 
from configparser import ConfigParser
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Border, Side, Alignment, Font
from openpyxl.styles.colors import RED,YELLOW, BLUE, BLACK,WHITE
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
#import sys

root = os.getcwd() #get the file folder in the current system
exceptionpath = root + '/worksheets/DailyReport/exception.csv'
aftermangementexception = root + '/worksheets/DailyReport/exception_afterdelduplication.csv'


def GetFile(filename):#the function is used SSL to download dailyfile from Iniffeon file server
     root = os.getcwd()
     path = root +'/worksheets/EmployeeDailyRecord/'
     ls = os.listdir(path)
     #print(ls)
     for l in ls:
          if l == filename:
                    p = path + '/'+l
                    os.remove(p)
     CONFIGFILE_FORSFTP = 'config.ini' #--read the configure file to get file from SFTP server
     config = ConfigParser() 
     config.read(CONFIGFILE_FORSFTP) 
     filepath = config.get('SFTP','filepath')#--
     root = os.getcwd()  
     sftpfilepath = filepath + filename
     localfilepath = path +filename
     transport = paramiko.Transport(('SFTPAP.Extra.infineon.com', 22))
     transport.connect(username='DAPLENELEXT-FTP', password='Siapmkd!@#$%^&*()')
     sftp = paramiko.SFTPClient.from_transport(transport)
     sftp.get(sftpfilepath, localfilepath)
     transport.close()


def MakeEmployeNumer():# The function is used for getting employeeID from NameList
     lines = []
     i = 0
     path = root + '/worksheets/EmployeeDoc/Name List.xlsx'
     df = pd.DataFrame(pd.read_excel(path))
     df = df[['Employee No.']]# Fitter other columns 
     rows = df.shape[0]
     for  i in range(0, rows):
           lines.append(df.iat[i, 0])
     return lines

def CheckDailyEmployInfo(dailyInfo):# Get every employee daily record 
    path = root +'/worksheets/EmployeeDailyRecord/'+dailyInfo
    with open(path ,"r") as data:
             employeDailyInfos = [line.strip() for line in data.readlines()]
             return employeDailyInfos

def AddEmployeInfo(dailyfilepath):# Input every employee records which in 'Name List'
     ename = ''
     one_employe_total  = []
     dailyinfo =  CheckDailyEmployInfo(dailyfilepath)
     employesname =  MakeEmployeNumer() 
     employesname_ = []
     dailyemployeRecords = []
     for ep in employesname:
          e = 'SIA'+ str(ep)
          employesname_.append(e)
     for da in dailyinfo:
                 new_da=da.strip()
                 for ep in employesname_:
                       if ep in new_da:
                           one_employe_total.append(da) 
                           ename = root +'/worksheets/EmployeeRecords/'+ ep +'.csv'
                           dailyemployeRecords.append(ep)
                           with open(ename,"a+") as f5:# Create employee files and write the employeee record
                                  for one in one_employe_total:
                                                   f5.writelines(one+'\n')
                                  ename = ''
                                  one_employe_total  = []
     return dailyemployeRecords#reaturn the CSV filename list

def FindEmployeNumer(kind):# got the IDL/DL employeer 
    lines = []
    i = 0
    path = root + '/worksheets/EmployeeDoc/Name List.xlsx'
    df = pd.DataFrame(pd.read_excel(path))# Use pandas dataframe to read data from excel file
    df = df[['Employee No.','IDL/DL']]
    if kind is 'DL':
       df = df.loc[df['IDL/DL']=='DL']
       rows = df.shape[0]
       for  i in range(0, rows):
            lines.append(df.iat[i,0])
    else:
        df = df.loc[df['IDL/DL']=='IDL']
        rows = df.shape[0]
        for  i in range(0, rows):
           lines.append(df.iat[i,0])
    return lines

def NumberConvertToString(employesname):#the filename for store data is named SIA + emplpyer.CSV,for chekcing must and SIA
     employesname_ = []
     for ep in employesname:
          e = 'SIA'+str(ep)
          employesname_.append(e)
     return employesname_

def ReadFile(filename,dateStriing): # According the employer store file  to find records which match the date
     targetRecords = []  
     with open(filename,"r") as employefile:
             lines = employefile.readlines()
             for line in lines:
                 new_line=line.strip()
                 if dateStriing in new_line:
                         targetRecords.append(line)
                         employefile.close() 
                                                       
     return targetRecords

def dataSpiltTime(records,fromtime,endtime,workstart,workend):#
     fromtime_1 = time.strptime(fromtime, "%Y%m%d%H:%M")#start to lunch time
     endtime_1 = time.strptime(endtime, "%Y%m%d%H:%M")#end of lunch time
     workstart = time.strptime(workstart, "%Y%m%d%H:%M")#start to work time
     workend = time.strptime(workend, "%Y%m%d%H:%M")#end of work time
     i = 0
     datas = []
     afterSpilt = []
     datas = records#get the record for which happend in the date
     for data in datas:
         e = data[0:16]#got the first to 16 chars for every record
         afterSpilt.append(e)
     n = len(afterSpilt)#
     index = []
     if (n>=2):# Algorithm principle：you can think list[1,2,3,4] 4 records for one person ,Use first one compare to Second one ,and Secondone to third one and so go 
          for i in range(0,n-1): 
               for  j in range(0,n-1-i):
                    InOrOut = afterSpilt[j]
                    InOrOut = InOrOut[0:4]
                    InOrOut_1 = afterSpilt[j+1]
                    InOrOut_1 = InOrOut_1[0:4]
                    if(InOrOut=='0220' and InOrOut_1=='0110'):
                         time_1 = afterSpilt[j]
                         time_1 = time_1[4:14]+':'+ time_1[14:16]
                         time_2 = afterSpilt[j+1]
                         time_2 = time_2[4:14]+':'+ time_2[14:16]
                         t1 = time.strptime(time_1, "%Y%m%d%H:%M")#第一次时间
                         t2 = time.strptime(time_2,"%Y%m%d%H:%M")#第二次时间
                         x = time.mktime(t2)-time.mktime(t1) #找到间隔时间大于10分钟的记录
                         y1 = time.mktime(t1)-time.mktime(fromtime_1)#第一次异常时间减去午饭开始时间
                         y2 = time.mktime(t1)-time.mktime(workstart)#第一次异常时间减去正常上班时间
                         y3 = time.mktime(t1)-time.mktime(endtime_1)#第一次异常时间减去午饭结束时间
                         y4 = time.mktime(t1)-time.mktime(workend)#第一次异常时间减去工作结束时间
                         y5 = time.mktime(t2)-time.mktime(fromtime_1)#第二次异常时间减去午饭开始时间
                         y6 = time.mktime(t2)-time.mktime(workstart)#第二次异常时间减去工作开始时间
                         y7 = time.mktime(t2)-time.mktime(endtime_1)#第二次异常时间减去午饭结束时间
                         y8 = time.mktime(t2)-time.mktime(workend)#第二次异常时间减去工作结束时间
                         if(y2 > 0 and y4 < 0 and y6> 0 and y8 < 0 and x > 600):
                         #if  y1 < 0 :
                              if(y3 > 0 and y7 > 0 ):
                                    index.append(j)
                              if(y1 < 0 and y5 < 0 ):
                                    index.append(j)
                              if(y1 > 0 and y7 < 0 and x > 2700):
                                    index.append(j)
                    elif(InOrOut=='0220' and InOrOut_1=='0220'):
                         time_1 = afterSpilt[j]
                         time_1 = time_1[4:14]+':'+ time_1[14:16]
                         time_2 = afterSpilt[j+1]
                         time_2 = time_2[4:14]+':'+ time_2[14:16]
                         t1 = time.strptime(time_1, "%Y%m%d%H:%M")#第一次时间
                         t2 = time.strptime(time_2,"%Y%m%d%H:%M")#第二次时间
                         x = time.mktime(t2)-time.mktime(t1) #找到间隔时间大于10分钟的记录
                         y1 = time.mktime(t1)-time.mktime(fromtime_1)#第一次异常时间减去午饭开始时间
                         y2 = time.mktime(t1)-time.mktime(workstart)#第一次异常时间减去正常上班时间
                         y3 = time.mktime(t1)-time.mktime(endtime_1)#第一次异常时间减去午饭结束时间
                         y4 = time.mktime(t1)-time.mktime(workend)#第一次异常时间减去工作结束时间
                         y5 = time.mktime(t2)-time.mktime(fromtime_1)#第二次异常时间减去午饭开始时间
                         y6 = time.mktime(t2)-time.mktime(workstart)#第二次异常时间减去工作开始时间
                         y7 = time.mktime(t2)-time.mktime(endtime_1)#第二次异常时间减去午饭结束时间
                         y8 = time.mktime(t2)-time.mktime(workend)#第二次异常时间减去工作结束时间
                         if(y2 > 0 and y4 < 0 and y6> 0 and y8 < 0 and x > 600):
                              if(y3 > 0 and y7 > 0 ):
                                        index.append(j)
                              if(y1 < 0 and y5 < 0 ):
                                        index.append(j)
                              if(y1 > 0 and y7 < 0 and x > 2700):
                                        index.append(j)
                    elif(InOrOut=='0110' and InOrOut_1=='0110'):
                         time_1 = afterSpilt[j]
                         time_1 = time_1[4:14]+':'+ time_1[14:16]
                         time_2 = afterSpilt[j+1]
                         time_2 = time_2[4:14]+':'+ time_2[14:16]
                         t1 = time.strptime(time_1, "%Y%m%d%H:%M")#第一次时间
                         t2 = time.strptime(time_2,"%Y%m%d%H:%M")#第二次时间
                         x = time.mktime(t2)-time.mktime(t1) #找到间隔时间大于10分钟的记录
                         y1 = time.mktime(t1)-time.mktime(fromtime_1)#第一次异常时间减去午饭开始时间
                         y2 = time.mktime(t1)-time.mktime(workstart)#第一次异常时间减去正常上班时间
                         y3 = time.mktime(t1)-time.mktime(endtime_1)#第一次异常时间减去午饭结束时间
                         y4 = time.mktime(t1)-time.mktime(workend)#第一次异常时间减去工作结束时间
                         y5 = time.mktime(t2)-time.mktime(fromtime_1)#第二次异常时间减去午饭开始时间
                         y6 = time.mktime(t2)-time.mktime(workstart)#第二次异常时间减去工作开始时间
                         y7 = time.mktime(t2)-time.mktime(endtime_1)#第二次异常时间减去午饭结束时间
                         y8 = time.mktime(t2)-time.mktime(workend)#第二次异常时间减去工作结束时间
                         if(y2 > 0 and y4 < 0 and y6> 0 and y8 < 0 and x > 600):
                              if(y3 > 0 and y7 > 0 ):
                                   index.append(j)
                              if(y1 < 0 and y5 < 0 ):
                                   index.append(j)
                              if(y1 > 0 and y7 < 0 and x > 2700):
                                   index.append(j)
     return index# return the location for the issue record in the list of records 
def  delstring(datas):#filtter some unuseful chars  for every records
     records = []
     for r in datas:
          e = r[0:24]
          records.append(e)
     return records

def  delduplicates():# Del duplicates to keep the data unique
     df = read_csv(exceptionpath)
     df = df.drop_duplicates()
     df.index = range(len(df))  
     df.to_csv(aftermangementexception,index=False)

def  SearchEmployeName():
     reader = []
     path = root + '/worksheets/EmployeeDoc/Name List.xlsx'
     df = pd.DataFrame(pd.read_excel(path))
     df = df[['Employee No.','English\nName']]
     csv_headers = ['Employee No.','EnglishName']
     df.to_csv(root + '/worksheets/EmployeeDoc/namelist.csv', header = csv_headers, index=False, mode='w',encoding='utf-8')
     with open(root + '/worksheets/EmployeeDoc/namelist.csv','r') as csvfile:
             lines = csvfile.readlines()
             for line in lines:
                    reader.append(line)
     return reader
def  CovertStrtoTime(times):# Convert the string datetiem to datetime format
     times = time.strptime(times, "%Y%m%d%H:%M")
     times = time.mktime(times)
     return times

def  exceptionFilter(filterFrom,filterEnd,excptionTimeFrom,exceptionTimeEnd):#For count how many affect times
      mins = 0 
      if    excptionTimeFrom < filterFrom and exceptionTimeEnd < filterFrom: 
                   mins = (exceptionTimeEnd - excptionTimeFrom)/60
      elif  excptionTimeFrom < filterFrom and exceptionTimeEnd <= filterEnd and exceptionTimeEnd >= filterFrom:
                   mins = (filterFrom - excptionTimeFrom)/60
      elif  excptionTimeFrom >= filterFrom and excptionTimeFrom < filterEnd and exceptionTimeEnd <= filterEnd:
                    t = (exceptionTimeEnd - excptionTimeFrom)/60
                    if t > 45:
                         mins = t
                    else:
                         mins = 0 
      elif  excptionTimeFrom > filterFrom  and excptionTimeFrom <= filterEnd  and exceptionTimeEnd > filterEnd:
                   mins = (exceptionTimeEnd - filterEnd)/60 
      elif  excptionTimeFrom >= filterEnd and exceptionTimeEnd >= filterEnd:
                   mins = (exceptionTimeEnd - excptionTimeFrom)/60
      return mins
                   
def  analyze(path):# the analyze every recodr during the date 
     lines = []
     adds = []
     adds = SearchEmployeName()
    # print(adds)
     i = 0
     df = read_csv(path)
     #print(df)
     rows = df.shape[0]
     for  i in range(0, rows):
            lines.append(df.iat[i,0])
     #print(lines)
     employes = []
     
     for  i in range(0, len(lines)-1,2):# use dict to store every record
          employee = { 'ID':'',
                       'Name':'',
                       'Date':'',
                       'From':'',
                       'End':'',
                       'Time(mins)':'',
                       'Type':''}#due to dict address you must everytime set a null dict
          col  = lines[i] 
          eid = col[19:24]
          #print(eid)
          tag = col
          employee['ID'] = eid
          for a in adds:
               if eid in a:
                    ename = a
                    ename = ename[6:len(adds)-2]
                    #print(ename)
          employee['Name'] = ename
          #employee['EmployeeRecords'] = col
          employee['Date'] = col[4:8]+'-'+col[8:10]+'-'+col[10:12]
          fromtime = col[12:14]+':'+ col[14:16]
          employee['From'] = fromtime
          col1 = lines[i+1]
          n = len(col)
          tag = col[24:n]
          endtime = col1[12:14]+':'+ col1[14:16]
          employee['End'] = endtime 
          m = CovertStrtoTime(lunchtimefrom) 
          n = CovertStrtoTime(lunchtimeto)
          outtime = CovertStrtoTime(col[4:14]+':'+ col[14:16])
          intime = CovertStrtoTime(col1[4:14]+':'+ col1[14:16])
          employee['Time(mins)'] = exceptionFilter(m ,n,outtime,intime)
          employee['Type'] = tag
          employes.append(employee)
     #employes.extend(extendrecords)
     return employes

def  CheckStartAndEnd(records,workstart,workend):
     workstart = time.strptime(workstart, "%Y%m%d%H:%M")#start to work time
     workend = time.strptime(workend, "%Y%m%d%H:%M")#end of work time
     employes = []
     adds = []
     adds = SearchEmployeName()
     #print(adds)
     datas = []
     datas_In = []
     datas_Out = []
     tag_1 = '0110'
     tag_2 = '0220'
     for r in records:
          if r not in  datas:
               datas.append(r)
     for data in datas:
          #e = data[0:16]#got the first to 16 chars for every record
          if   tag_1  in data:
               datas_In.append(data)
          elif tag_2  in data:
               datas_Out.append(data)
     #print(afterSpilt)
     n1  = len(datas_In)#
     n2  = len(datas_Out)
     #In = datas_In[0]
     #In = InOrOut[0:4]
     if(n1>0):
          time_In = datas_In[0]
          time_In = time_In[4:14]+':'+ time_In[14:16]
          t_In = time.strptime(time_In, "%Y%m%d%H:%M")#第一次进门刷卡时间
          y1 = time.mktime(t_In)-time.mktime(workstart)#第一次刷卡进入时间减去正常上班时间
          y2 = time.mktime(t_In)-time.mktime(workend)#第一次刷卡进入时间减去正常下班时间
          if(y1>0 and y2<0 and y1>180):
               employee = { 'ID':'',
                       'Name':'',
                       'Date':'',
                       'From':'',
                       'End':'',
                       'Time(mins)':'',
                       'Type':''}#due to dict address you must everytime set a null dict
               col  = datas_In[0]
               eid = col[19:24]
               employee['ID'] = eid
               for a in adds:
                    if eid in a:
                         ename = a
                         ename = ename[6:len(adds)-2]
               employee['Name'] = ename
               employee['Date'] = col[4:8]+'-'+col[8:10]+'-'+col[10:12]
               fromtime = '09:00'
               employee['From'] = fromtime
               endtime = col[12:14]+':'+ col[14:16]
               employee['End'] = endtime 
               employee['Time(mins)'] = y1/60
               employee['Type'] = 'BeLate'
               employes.append(employee)
     if(n2>0):
          time_Out = datas_Out[n2-1]
          time_Out = time_Out[4:14]+':'+ time_Out[14:16]
          t_Out = time.strptime(time_Out, "%Y%m%d%H:%M")#最后一次出门刷卡时间
          y3 = time.mktime(t_Out)-time.mktime(workstart)#最后一次刷卡出去时间减去正常上班时间
          y4 = time.mktime(workend)-time.mktime(t_Out)#最后一次刷卡出去时间减去正常下班时间
          if(y3>0 and y4>0 and y4>180):
               employee = { 'ID':'',
                       'Name':'',
                       'Date':'',
                       'From':'',
                       'End':'',
                       'Time(mins)':'',
                       'Type':''}#due to dict address you must everytime set a null dict
               col  = datas_Out[n2-1]
               eid = col[19:24]
               employee['ID'] = eid
               for a in adds:
                    if eid in a:
                         ename = a
                         ename = ename[6:len(adds)-2]
               employee['Name'] = ename
               employee['Date'] = col[4:8]+'-'+col[8:10]+'-'+col[10:12]
               fromtime =  col[12:14]+':'+ col[14:16]
               employee['From'] = fromtime
               endtime = '17:30'
               employee['End'] = endtime 
               employee['Time(mins)'] = y4/60
               employee['Type'] = 'LeaveEearly'
               employes.append(employee)
     return employes      
def  SaveAsExcelReport(filename,date,employes):
      if(employes):
          csv_headers = [ 'Date',
                     'End',
                     'From',
                     'ID',
                     'Name',
                     'Time(mins)',
                     'Type'                
           ]
          df = pd.DataFrame(employes)#use pandas dataframe to got all dicts
          df.index = range(len(df))  # got the index
          df.to_csv(filename, header=csv_headers, index=False, mode='a+')#save to csv file
          #df.to_csv(filename, index=False, mode='a+')#save to csv file
          df = pd.read_csv(filename)# read the CSV file 
          print(df)
          cols=[ 'ID','Name','Date','From','End','Time(mins)','Type']#setting excel file column
          df=df.loc[:,cols]#keep the data display in the excel file follow the header format
          df = df.loc[df['Time(mins)']!= 0]# filtter the records which affect time eq = 0
          #df = df.loc[df['ID']!=20182]
          df.to_excel(root + '/worksheets/ForHR/exception_records_'+ date +'.xlsx',index=False, sheet_name='Data')
          return True
      else:    
          
          return False
     

def  GotExcelRows(date):# got the total coulumns for the current excel file
     path = root + '/worksheets/ForHR/exception_records_'+date+'.xlsx'
     df = pd.DataFrame(pd.read_excel(path))
     rows = df.shape[0] 
     return rows

def  DrewChart(path,chartlocation):# drew the char  in the report excel file
     location =  chartlocation + 4
     location_1 ='A' + str(location)
     wb = load_workbook(path)
     ws = wb.active
     chart1 = BarChart()
     chart1.type = "col"
     chart1.style = 10
     chart1.title = "Exception Chart"
     chart1.y_axis.title = 'Time(times)'
     chart1.x_axis.title = 'Name'
     data = Reference(ws, min_col=6, min_row=1, max_row=14)#setting X-RAY
     cats = Reference(ws, min_col=2, min_row=2, max_row=14)#setting  Y-RAY
     chart1.add_data(data, titles_from_data=True)
     chart1.set_categories(cats)
     chart1.shape = 4
     ws.add_chart(chart1, location_1)
     wb.save(path)

def  setExcelFormat(filename):#setting excel format
     alignment=Alignment(horizontal='center',vertical='center')#--Setting Excel Format
     thin = Side(border_style="thin",color=BLACK)
     border = Border(top=thin, left=thin, right=thin, bottom=thin)
     row_title_font = Font(name='Times New Roman', size=12, bold=True, color=WHITE)#setting header format
     row_title_fill = PatternFill(fill_type='solid',fgColor=BLUE)
     content_font = Font(name='Times New Roman', size=10, bold=False,color=BLACK)#setting the cell format withou the header
     content_fill = PatternFill(fill_type='solid',fgColor=WHITE)#Setting Excel Format--
     wb = openpyxl.load_workbook(filename)
     #ws = wb.get_sheet_by_name('data')
     ws = wb['Data']
     ws.row_dimensions[1].height = 20
     ws.column_dimensions['B'].width = 16.0
     ws.column_dimensions['F'].width = 12.0
     for row in ws.rows:
          for cell in row:
                cell.alignment = alignment
                if cell.row == 1:
                      cell.border = border
                      cell.font = row_title_font
                      cell.fill = row_title_fill
                else:
                      cell.border = border
                      cell.font = content_font
                      cell.fill = content_fill
    
     wb.save(filename)

def  delfile():     #del file
     path = root + '/worksheets/DailyReport/exception_afterdelduplication.csv'  
     os.remove(path)
     os.remove(exceptionpath)

def  delfile_1(path):# del all csv file for store daily records,it is very import otherwise may affect the data
     ls = os.listdir(path)
     #print(ls)
     for l in ls:
          p = path + '/'+l
          os.remove(p)

def  SendEmail(attFile,date):
     CONFIGFILE_FORSendEmail = 'emailconfig.ini' 
     config = ConfigParser() 
     config.read(CONFIGFILE_FORSendEmail) 
     Sender = config.get('EmailSender','Sender') 
     Password = config.get('EmailSenderPassword','Password') 
     #Receiver = config.get('EmailReceiver','Receiver') 
     SMTPServer =  config.get('SMTPServer','Server') 
     SMTPServerPort = config.get('SMTPServerPort','ServerPort')
     mail_user = Sender
     mail_pass = Password 
     #mailto_list = Receiver
     mail_host= SMTPServer
     port = SMTPServerPort
     mailto_list = []
     with open(root + '/worksheets/Email/emailaddresss.txt','r') as csvfile:
             lines = csvfile.read().splitlines()
             for  line in lines:
                    mailto_list.append(line)
     msg = MIMEMultipart()
     msg['Subject'] = "EmployeeDailyRecord" 
     msg['From'] = mail_user
     msg['To'] =  ','.join(mailto_list)
     strstr="Attachment is the abnormal record of employee attendance yesterday,Please find it "  #文字内容
     att = MIMEText(strstr,'plain','utf-8')
     msg.attach(att)
     att = MIMEApplication(open(attFile,'rb').read())  #你要发送的附件地址
     att.add_header('Content-Disposition', 'attachment', filename='exception_records_'+date+'.xlsx') #filename可随意取名
     msg.attach(att)
     server = smtplib.SMTP(mail_host,port)
     server.starttls()
     server.login(mail_user,mail_pass) 
     #m = ['348532554@qq.com','rey.wang@siapm.com.cn']
     server.sendmail(mail_user,mailto_list, msg.as_string())    #发送
     server.close()  #关闭

def  SendEmailForNormal():
     CONFIGFILE_FORSendEmail = 'emailconfig.ini' 
     config = ConfigParser() 
     config.read(CONFIGFILE_FORSendEmail) 
     Sender = config.get('EmailSender','Sender') 
     Password = config.get('EmailSenderPassword','Password') 
     #Receiver = config.get('EmailReceiver','Receiver') 
     SMTPServer =  config.get('SMTPServer','Server') 
     SMTPServerPort = config.get('SMTPServerPort','ServerPort')
     mail_user = Sender
     mail_pass = Password 
     #mailto_list = Receiver
     mail_host= SMTPServer
     port = SMTPServerPort
     mailto_list = []
     with open(root + '/worksheets/Email/emailaddresss.txt','r') as csvfile:
             lines = csvfile.read().splitlines()
             for  line in lines:
                    mailto_list.append(line)
     msg = MIMEMultipart()
     msg['Subject'] = "EmployeeDailyRecord" 
     msg['From'] = mail_user
     msg['To'] =  ','.join(mailto_list)
     strstr="There is no exception record yesterday!"  #文字内容
     att = MIMEText(strstr,'plain','utf-8')
     msg.attach(att)
     #att = MIMEApplication(open(attFile,'rb').read())  #你要发送的附件地址
     #att.add_header('Content-Disposition', 'attachment', filename="DailyReport.xlsx") #filename可随意取名
     #msg.attach(att)
     server = smtplib.SMTP(mail_host,port)
     server.starttls()
     server.login(mail_user,mail_pass)   #登录
     server.sendmail(mail_user, mailto_list, msg.as_string())    #发送
     server.close()  #关闭
if __name__ == '__main__':
     print('Start the Program now, Please Wait until all steps finished!')
     now_date = datetime.datetime.now()# got current date
     currentdate = datetime.datetime.now().strftime('%Y%m%d')
     yes_date = (now_date -3*Day()).strftime('%Y%m%d')# got yesterday date
     #daily =  currentdate +'.txt'# this file name is store in SFTP server
     daily =   '20190907.txt'# this file name is store in SFTP server
     print(' 1: Start to download the data file from SFTP Server!')
     GetFile(daily)#download the file from SFTP server
     print(' 2: Download Data file finished!')
     print(' 3: Start to get the record from data file!')
     dailyExitRecords = AddEmployeInfo(daily)# got the filename list
     print(' 4: Create the employe record file successfully!')
     print(' 5: Start to  check the EmployeeRecordFile ....')
     kind = 'IDL'
     employesname =  FindEmployeNumer(kind)
     employesname_1 =  NumberConvertToString(employesname)
     employesname_2 = []
     delemployesname = []
     for da in employesname_1:# from all IDL emplpyeers fo find which one has record in daily
                 new_da=da.strip()
                 for ep in dailyExitRecords:
                                   if ep in new_da:
                                             employesname_2.append(da)   
     #print(employesname_2)
     for e in employesname_2:
               if e not in  delemployesname:
                         delemployesname.append(e)
     #print(delemployesname)
     lunchtimefrom = yes_date +'11:30' 
     lunchtimeto = yes_date +'13:00'
     limitStart =  yes_date +'09:00'
     limitEnd = yes_date +'17:30'
     fromtime_1 = time.strptime(lunchtimefrom, "%Y%m%d%H:%M")#午饭开始时间转码
     endtime_1 = time.strptime(lunchtimeto, "%Y%m%d%H:%M")#午饭结束时间
     records = []
     #print(records)
     result_step1 = []
     result_step2 = []
     exception_1  = []
     exception_2  = []
     records = []
     with open(exceptionpath,"w") as f:
               f.writelines('EmployeeRecord'+'\n')
               f.close()
     for e in  delemployesname:
               filename =root +'/worksheets/EmployeeRecords/'+ e +'.csv'
               records = ReadFile(filename,yes_date)
               records = delstring(records)
               #print(records)
               index = dataSpiltTime(records,lunchtimefrom,lunchtimeto,limitStart,limitEnd)
               #print(index)
               lst = []
               for i in index:                    #循环list里的每一个元素
                    if  i not in lst:            #判断元素是否存在新列表中，不存在则添加，存在则跳过，以此去重
                              lst.append(i)
               for i in lst:
                    InOrOut = records[i][0:16]#
                    InOrOut = InOrOut[0:4]
                    InOrOut_1 = records[i+1][0:16]
                    InOrOut_1 = InOrOut_1[0:4]
                    time_1 = records[i][0:16]
                    time_1 = time_1[4:14]+':'+ time_1[14:16]
                    time_2 = records[i+1][0:16]
                    time_2 = time_2[4:14]+':'+ time_2[14:16]
                    t1 = time.strptime(time_1, "%Y%m%d%H:%M")#第一次时间
                    t2 = time.strptime(time_2,"%Y%m%d%H:%M")#第二次时间
                    y1 = time.mktime(t1)-time.mktime(fromtime_1)
                    y2 = time.mktime(t2)-time.mktime(endtime_1)
                    y3 = time.mktime(t1)-time.mktime(endtime_1)
                    y4 = time.mktime(t2)-time.mktime(fromtime_1)
                    if  (InOrOut=='0220'and InOrOut_1=='0110' and  y1 < 0 and y4 <0):
                                   result_step1.append(records[i]+'OUT/IN')
                                   result_step1.append(records[i+1]+'OUT/IN')
                    elif(InOrOut=='0220'and InOrOut_1=='0110' and  y2 > 0 and y3 >0):
                                   result_step1.append(records[i]+'OUT/IN')
                                   result_step1.append(records[i+1]+'OUT/IN')
                    elif(InOrOut=='0110'and InOrOut_1=='0110' and  y1 < 0 and y4 <0):
                                   result_step1.append(records[i]+'IN/IN')
                                   result_step1.append(records[i+1]+'IN/IN')
                    elif(InOrOut=='0110'and InOrOut_1=='0110' and  y2 > 0 and y3 >0):
                                   result_step1.append(records[i]+'IN/IN')
                                   result_step1.append(records[i+1]+'IN/IN')
                    else:
                                   result_step1.append(records[i]+'BreakTime')
                                   result_step1.append(records[i+1]+'BreakTime')
                    with open(exceptionpath,"a+") as f:
                              for e in result_step1:
                                   f.writelines(e+'\n')
  
     print('6: Store exception records in the CSV file successfully!')   
     print('7: Start to delete duplicate exception records  ....') 
     delduplicates()
     print('8: Delete duplicate exception records that may exist successfully!')
     print('9: Start to analyze the ExceptionRecordFile ....') 
     check = []
     for e in  delemployesname:
               filename =root +'/worksheets/EmployeeRecords/'+ e +'.csv'
               records = ReadFile(filename,yes_date)
               records = delstring(records)
               res1 = CheckStartAndEnd(records,limitStart,limitEnd)
               #print(results_1)
               check.extend(res1)
     #print(check)
     path = root + '/worksheets/DailyReport/exception_afterdelduplication.csv'
     path_1 = root + '/worksheets/ForHR/exception_records_'+yes_date+'.xlsx'
     path_2 = root + '/worksheets/EmployeeRecords'
     filename = root + '/worksheets/DailyReport/exception_records_'+yes_date+'.csv'
     #analyze(path,filename,yes_date)
     results = analyze(path)
     results.extend(check)
     #print(results)
     #print(check)
     #results.extend(check)
     #print(results)
     res = SaveAsExcelReport(filename,yes_date,results)
     print('10: Analyze the record in the exception file Successfully!')
     if(res):
          print('11: Start to setting the report format ....') 
          setExcelFormat(path_1)
          print('12: Setting the report format Successfully!')
          print('13: Start to delete the unnecessary data processing files ....') 
          delfile()
          delfile_1(path_2)
          print('14: Delete the unnecessary data processing files')
          print('15: Start to create chart for the report ....') 
          rows = GotExcelRows(yes_date)
          DrewChart(path_1,rows)
          print('16: Create Chart Successfully!')
          print('17: Start to send the current report by email ....') 
          SendEmail(path_1,yes_date)
          print('18: Send the current report by email successfully!')
          print('19: Start to close the program.....')
          print('20: All Task Finshed!')
     else:
          delfile()
          delfile_1(path_2)
          print('Alarm: No Exception records for the date and no report created!')
          SendEmailForNormal()
     